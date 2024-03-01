# -*- coding: utf-8 -*-

from data.base_data import BaseData
from data.util import data_utils
from data.util import file_utils
from . import excel_utils

import copy
import openpyxl
import os
import xlrd
import xlsxwriter


class ExcelData(BaseData):
    def __init__(self, filename):
        super().__init__(os.path.split(filename)[0])
        self.filename = filename

    def read(self):
        if self.filename.endswith('.xls'):
            self.read_xls()
        else:
            self.read_xlsx()

    def read_xls(self):
        schema = self.get_schema()

        workbook = xlrd.open_workbook(self.filename)

        sheet_dict = {}
        for sheet in workbook.sheets():
            table_name = schema.get_replace_table_name(sheet.name)
            if schema.is_valid() and not schema.has_table(table_name):
                continue

            sheet_dict[table_name] = sheet

        table_names = schema.get_table_names()

        for table_name in table_names:
            if table_name not in sheet_dict:
                columns = schema.get_columns(table_name)
                self.set_columns(table_name, columns)
                continue

            columns = None

            sheet = sheet_dict[table_name]
            for row_idx in range(sheet.nrows):
                if row_idx == 0:
                    columns = sheet.row_values(row_idx)
                    columns = [schema.get_replace_column_name(sheet.name, c) for c in columns]
                    self.set_columns(table_name, columns)
                else:
                    old_values = sheet.row_values(row_idx)
                    new_values = []
                    for col_idx in range(len(old_values)):
                        cell_type = sheet.cell_type(row_idx, col_idx)
                        if cell_type == excel_utils.XL_CELL_DATE:
                            cell_type = excel_utils.XL_CELL_TIMESTAMP

                        if columns is not None:
                            column_info = self.get_column_info(table_name, columns[col_idx])
                            if column_info is None:
                                data_type = data_utils.get_data_type(columns[col_idx], old_values[col_idx])
                                cell_type = excel_utils.to_cell_type(data_type)
                            else:
                                cell_type = excel_utils.to_cell_type(column_info['type'])

                        new_values += [excel_utils.to_value(sheet, cell_type, old_values[col_idx])]

                    self.add_data(table_name, new_values)

    def read_xlsx(self):
        schema = self.get_schema()

        workbook = openpyxl.load_workbook(self.filename)

        sheet_dict = {}

        sheet_names = workbook.get_sheet_names()
        for sheet_name in sheet_names:
            sheet = workbook.get_sheet_by_name(sheet_name)
            table_name = schema.get_replace_table_name(sheet_name)
            if schema.is_valid() and not schema.has_table(table_name):
                continue

            sheet_dict[table_name] = (sheet_name, sheet)

        table_names = schema.get_table_names()

        for table_name in table_names:
            if table_name not in sheet_dict:
                columns = schema.get_columns(table_name)
                self.set_columns(table_name, columns)
                continue

            columns = None

            sheet_name, sheet = sheet_dict[table_name]
            for i, row in enumerate(sheet):
                if i == 0:
                    columns = [col.value for col in row]
                    columns = [schema.get_replace_column_name(sheet_name, c) for c in columns]
                    self.set_columns(table_name, columns)
                else:
                    old_values = [col.value for col in row]
                    new_values = []
                    for col_idx in range(len(old_values)):
                        cell_type = None

                        if columns is not None:
                            column_info = self.get_column_info(table_name, columns[col_idx])
                            if column_info is None:
                                data_type = data_utils.get_data_type(columns[col_idx], old_values[col_idx])
                                cell_type = excel_utils.to_cell_type(data_type)
                            else:
                                cell_type = excel_utils.to_cell_type(column_info['type'])

                        new_values += [excel_utils.to_value(sheet, cell_type, old_values[col_idx])]

                    self.add_data(table_name, new_values)

    def write(self):
        new_filename = self.filename
        if new_filename.endswith('.xls'):
            new_filename = new_filename.replace('.xls', '.xlsx')

        workbook = None

        dest = self.get_dest()
        if dest is not None:
            workbook = xlsxwriter.Workbook(os.path.join(dest, os.path.split(new_filename)[-1]))
        else:
            workbook = xlsxwriter.Workbook(file_utils.new_filename(new_filename))

        format_header = workbook.add_format({
            'bold': True,
            'font_color': '#FFFFFF',
            'bg_color': '#404040',
            'top': 1,
            'bottom': 1,
            'left': 1,
            'right': 1,
            'top_color': '#666666',
            'bottom_color': '#666666',
            'left_color': '#666666',
            'right_color': '#666666',
            'align': 'left'
        })

        default_dict = {
            'bold': False,
            'font_color': '#000000',
            'bg_color': '#FFFFFF',
            'top': 1,
            'bottom': 1,
            'left': 1,
            'right': 1,
            'top_color': '#666666',
            'bottom_color': '#666666',
            'left_color': '#666666',
            'right_color': '#666666',
            'align': 'left'
        }

        date_dict = copy.deepcopy(default_dict)
        date_dict['num_format'] = 'yyyy-mm-dd'
        date_dict['align'] = 'center'

        time_dict = copy.deepcopy(default_dict)
        time_dict['num_format'] = 'hh-mm-ss'
        time_dict['align'] = 'center'

        timestamp_dict = copy.deepcopy(default_dict)
        timestamp_dict['num_format'] = 'yyyy-mm-dd hh-mm-ss'
        timestamp_dict['align'] = 'center'

        boolean_dict = copy.deepcopy(default_dict)
        boolean_dict['align'] = 'center'

        number_dict = copy.deepcopy(default_dict)
        number_dict['align'] = 'right'

        integer_dict = copy.deepcopy(default_dict)
        integer_dict['align'] = 'right'

        format_default = workbook.add_format(default_dict)
        format_date = workbook.add_format(date_dict)
        format_time = workbook.add_format(time_dict)
        format_timestamp = workbook.add_format(timestamp_dict)
        format_boolean = workbook.add_format(boolean_dict)
        format_number = workbook.add_format(number_dict)
        format_integer = workbook.add_format(integer_dict)

        table_names = self.get_schema().get_table_names()
        for table_name in table_names:
            data = self.get_data(table_name)

            worksheet = workbook.add_worksheet(table_name)
            worksheet.freeze_panes(1, 0)
            worksheet.hide_gridlines(2)

            if len(data) == 0:
                worksheet.hide()

            column_names = []

            columns = self.get_columns(table_name)
            if columns is None:
                continue

            row_idx = 0
            for col_idx, value in enumerate(columns):
                column_names += [value]
                worksheet.write(row_idx, col_idx, value, format_header)

            for row_data in data:
                row_idx += 1
                for col_idx, value in enumerate(row_data):
                    cell_format = format_default
                    if col_idx < len(column_names):
                        column_name = column_names[col_idx]
                        column_info = self.get_column_info(table_name, column_name)

                        column_type = 'string'
                        if column_info is None:
                            column_type = data_utils.get_data_type(column_name, value)
                        else:
                            column_type = column_info['type']

                        if column_type == 'date':
                            cell_format = format_date
                        elif column_type == 'time':
                            cell_format = format_time
                        elif column_type == 'timestamp':
                            cell_format = format_timestamp
                        elif column_type == 'boolean':
                            cell_format = format_boolean
                        elif column_type == 'number':
                            cell_format = format_number
                        elif column_type == 'integer':
                            cell_format = format_integer

                    worksheet.write(row_idx, col_idx, value, cell_format)

            worksheet.autofit()

        workbook.close()
